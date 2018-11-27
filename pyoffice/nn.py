from openpyxl import load_workbook
import re

      #通过索引顺序获取
wb = load_workbook(filename="dd.xlsx")
 
#获取当前活跃的worksheet,默认就是第一个worksheet
#ws = wb.active
 
#当然也可以使用下面的方法
 
#获取所有表格(worksheet)的名字
sheets = wb.get_sheet_names()
#第一个表格的名称
sheet_first = sheets[0]
#获取特定的worksheet
ws = wb.get_sheet_by_name(sheet_first)
for i in range(1,1082):
	s=ws.cell(row=i, column=5).value
	print(s)
	m = re.match(".*\((.*)\).*", str(s))
	if m==None:
		print(i)
	else:
		print(m.group(1))
        # data.write(i,5,m.group(1)) 
		ws.cell(row=i, column=6).value = m.group(1)
# #保存文件  
# file.save('tt.xlsx') 
wb.save(filename="a.xlsx")